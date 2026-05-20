from io import BytesIO
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

class export_excel:

    def export_excel(self):

        queryset = self.filter_queryset(self.get_queryset())

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Users"

        PRIMARY_COLOR = "1F4E78"
        SECONDARY_COLOR = "D9EAF7"
        HEADER_TEXT_COLOR = "FFFFFF"

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        worksheet.merge_cells("A1:L1")

        worksheet["A1"] = "USERS EXPORT REPORT"

        worksheet["A1"].font = Font(
            bold=True,
            size=18,
            color=HEADER_TEXT_COLOR,
        )

        worksheet["A1"].fill = PatternFill(
            start_color=PRIMARY_COLOR,
            end_color=PRIMARY_COLOR,
            fill_type="solid",
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 30

        request = self.request

        search_value = request.query_params.get(
            "search",
            "None",
        )

        ordering_value = request.query_params.get(
            "ordering",
            "Default",
        )

        filters = []

        for field in [
            "gender",
            "country",
            "created_at",
            "is_active",
            "is_admin",
        ]:

            value = request.query_params.get(field)

            if value:

                filters.append(f"{field}: {value}")

        filters_text = ", ".join(filters) if filters else "None"

        worksheet["A3"] = "Search:"
        worksheet["B3"] = search_value

        worksheet["A4"] = "Ordering:"
        worksheet["B4"] = ordering_value

        worksheet["A5"] = "Filters:"
        worksheet["B5"] = filters_text

        for row in range(3, 6):

            worksheet[f"A{row}"].font = Font(bold=True)

        headers = [
            "ID",
            "Email",
            "First Name",
            "Last Name",
            "Phone Number",
            "Gender",
            "Avatar",
            "Country",
            "Age",
            "Birth Date",
            "Is Active",
            "Is Admin",
            "Created At",
            "Updated At",
        ]

        header_row = 8

        for column_number, header in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color=HEADER_TEXT_COLOR,
            )

            cell.fill = PatternFill(
                start_color=PRIMARY_COLOR,
                end_color=PRIMARY_COLOR,
                fill_type="solid",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.border = thin_border

        data_row = 9

        for user in queryset:

            row_data = [
                user.id,
                user.email,
                user.first_name,
                user.last_name,
                user.phone_number,
                user.gender,
                (request.build_absolute_uri(user.avatar.url) if user.avatar else ""),
                user.country,
                user.age,
                (user.birth_date.strftime("%Y-%m-%d") if user.birth_date else ""),
                user.is_active,
                user.is_admin,
                (
                    user.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if user.created_at
                    else ""
                ),
                (
                    user.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                    if user.updated_at
                    else ""
                ),
            ]

            for column_number, value in enumerate(
                row_data,
                start=1,
            ):

                cell = worksheet.cell(
                    row=data_row,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border

                if data_row % 2 == 0:

                    cell.fill = PatternFill(
                        start_color=SECONDARY_COLOR,
                        end_color=SECONDARY_COLOR,
                        fill_type="solid",
                    )

            data_row += 1

        for column_cells in worksheet.columns:

            column_length = max(len(str(cell.value or "")) for cell in column_cells)

            adjusted_width = min(
                column_length + 5,
                50,
            )

            column_letter = get_column_letter(column_cells[0].column)

            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.freeze_panes = "A9"

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = 'attachment; filename="users_report.xlsx"'

        return response
