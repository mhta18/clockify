import os
import openpyxl
from celery import shared_task
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
from timetracking.models import TimeLog
from django.db.models import (
    Sum,
    ExpressionWrapper,
    F,
    fields,
    FloatField,
    Count,
    Avg,
    Q,
)
from django.conf import settings
from django.db.models.functions import Coalesce
from django.core.mail import EmailMessage
from django.utils import timezone
from users.models import User
from contracts.models import FreelancerContract, EmployerContract


@shared_task(name="reports.tasks.generate_platform_excel_report")
def generate_platform_excel_report(admin_email):
    completed_logs = TimeLog.objects.filter(end_time__isnull=False)
    duration_expression = ExpressionWrapper(
        F("end_time") - F("start_time"), output_field=fields.DurationField()
    )

    global_stats = completed_logs.annotate(
        calculated_duration=duration_expression
    ).aggregate(
        total_logs_payout=Coalesce(Sum("payment"), Decimal("0.00")),
        total_duration_timedelta=Sum("calculated_duration"),
    )

    global_seconds = (
        global_stats["total_duration_timedelta"].total_seconds()
        if global_stats["total_duration_timedelta"]
        else 0
    )
    global_hours = round(Decimal(global_seconds) / Decimal("3600.00"), 2)

    freelancer_logs = completed_logs.filter(
        user__contract__freelancercontract__isnull=False
    ).annotate(calculated_duration=duration_expression)
    average_age = User.objects.aggregate(
        average_age=Coalesce(
            ExpressionWrapper(
                Avg("age", filter=Q(age__isnull=False)), output_field=FloatField()
            ),
            0.0,
        )
    )["average_age"]

    global_status = User.objects.aggregate(
        total_users=Count("id"),
        total_men=Count("id", filter=Q(gender="male")),
        total_women=Count("id", filter=Q(gender="female")),
        total_unspecified=Count("id", filter=Q(gender="other")),
    )

    freelancer_stats = freelancer_logs.aggregate(
        total_payout=Coalesce(Sum("payment"), Decimal("0.00")),
        total_time=Sum("calculated_duration"),
    )
    fl_demographics = FreelancerContract.objects.aggregate(
        men_count=Count("user", distinct=True, filter=Q(user__gender="male")),
        women_count=Count("user", distinct=True, filter=Q(user__gender="female")),
        other_count=Count("user", distinct=True, filter=Q(user__gender="other")),
    )

    fl_total_hours = round(
        Decimal(
            freelancer_stats["total_time"].total_seconds()
            if freelancer_stats["total_time"]
            else 0
        )
        / Decimal("3600.00"),
        2,
    )

    employer_logs = completed_logs.filter(
        user__contract__employercontract__isnull=False
    ).annotate(calculated_duration=duration_expression)
    employer_stats = employer_logs.aggregate(
        total_payout=Coalesce(Sum("payment"), Decimal("0.00")),
        total_time=Sum("calculated_duration"),
    )
    emp_demographics = EmployerContract.objects.aggregate(
        men_count=Count("user", distinct=True, filter=Q(user__gender="male")),
        women_count=Count("user", distinct=True, filter=Q(user__gender="female")),
        other_count=Count("user", distinct=True, filter=Q(user__gender="other")),
    )

    emp_total_hours = round(
        Decimal(
            employer_stats["total_time"].total_seconds()
            if employer_stats["total_time"]
            else 0
        )
        / Decimal("3600.00"),
        2,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Platform Summary"
    ws.views.sheetView[0].showGridLines = True

    # Define clean styling sheets
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_section = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)

    fill_header = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Add Title
    ws["B2"] = f"PLATFORM ADMINISTRATIVE SUMMARY REPORT FOR {admin_email}"
    ws["B2"].font = font_title
    current_row = 4

    def write_table_block(title, data_dict):
        nonlocal current_row

        ws.merge_cells(
            start_row=current_row, start_column=2, end_row=current_row, end_column=3
        )
        header_cell = ws.cell(row=current_row, column=2, value=title)
        header_cell.font = font_section
        header_cell.fill = fill_header
        header_cell.alignment = align_left
        current_row += 1

        # Data Rows
        for index, (metric_name, metric_value) in enumerate(data_dict.items()):
            c_label = ws.cell(row=current_row, column=2, value=str(metric_name))
            c_val = ws.cell(row=current_row, column=3, value=metric_value)

            c_label.font = font_regular
            c_label.alignment = align_left
            c_val.font = font_bold
            c_val.alignment = align_right

            # Format numerical strings vs currencies nicely
            if "payout" in metric_name.lower() or "cost" in metric_name.lower():
                c_val.number_format = "$#,##0.00"
            elif "hours" in metric_name.lower():
                c_val.number_format = "#,##0.00"
            elif isinstance(metric_value, int):
                c_val.number_format = "#,##0"

            # Zebra striping
            if index % 2 == 1:
                c_label.fill = fill_zebra
                c_val.fill = fill_zebra

            current_row += 1
        current_row += 2

    write_table_block(
        "1. Platform-Wide Totals",
        {
            "Total Completed Hours": global_hours,
            "Total Processed Payouts": global_stats["total_logs_payout"],
        },
    )

    write_table_block(
        "2. Global User Overview",
        {
            "Total Registered Accounts": global_status["total_users"],
            "All Men Registered": global_status["total_men"],
            "All Women Registered": global_status["total_women"],
            "Unspecified Gender": global_status["total_unspecified"],
            "Average User Age": round(average_age, 2) if average_age else 0.0,
        },
    )

    write_table_block(
        "3. Freelancer Operational Metrics",
        {
            "Total Hours Tracked": fl_total_hours,
            "Total Payout Processed": freelancer_stats["total_payout"],
            "Total Active Contracts": fl_demographics["men_count"]
            + fl_demographics["women_count"],
            "Men Count": fl_demographics["men_count"],
            "Women Count": fl_demographics["women_count"],
            "Other/Unspecified Count": fl_demographics["other_count"],
        },
    )

    write_table_block(
        "4. Employer Operational Metrics",
        {
            "Total Hours Tracked": emp_total_hours,
            "Total Costs Accumulated": employer_stats["total_payout"],
            "Total Active Contracts": emp_demographics["men_count"]
            + emp_demographics["women_count"],
            "Men Count": emp_demographics["men_count"],
            "Women Count": emp_demographics["women_count"],
            "Other/Unspecified Count": fl_demographics["other_count"],
        },
    )

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    timestamp = timezone.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"Platform_Administrative_Report_{timestamp}.xlsx"
    temp_dir = os.path.join(settings.MEDIA_ROOT, "temp_excel_exports")

    os.makedirs(temp_dir, exist_ok=True)
    file_path = os.path.join(temp_dir, filename)
    wb.save(file_path)

    try:
        email_msg = EmailMessage(
            subject=f"Administrative Platform Summary Report - {timezone.now().date()}",
            body="Hello Admin,\n\nThe platform-wide summary report has finished processing. Please check the attached spreadsheet for full metrics breakdown.",
            from_email=settings.EMAIL_HOST_USER,
            to=[admin_email],
        )
        email_msg.attach_file(file_path)
        email_msg.send(fail_silently=False)

        return f"Report cleanly sent to {admin_email}"
    except Exception as e:
        raise e
